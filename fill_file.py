from datascience import *
from openpyxl import *
import datetime
from tkinter import *
from tkinter import messagebox
import getpass
from tkinter.ttk import Combobox

test1 = Table.read_table("October.csv")
root = Tk()
root.title("Fill File")
root.geometry("640x200+0+0")
#label1 = Label(root, text = "Enter the site name ").place(x = 0, y = 0)
labelTop = Label(root,
                    text = "Select site name")
labelTop.grid(column=0, row=0)
ls = test1.group("Agency").column("Agency").tolist()
comboExample = Combobox(root, values= ls)

comboExample.grid(column=5, row=0)


label2 = Label(root, text = "Enter the .csv file name (exclude '.csv' from entry) ").place(x = 0, y = 30)
label3 = Label(root, text = "Enter the month (number) ").place(x = 0, y = 60)
label4 = Label(root, text = "Enter the year (4-digit number) ").place(x = 0, y = 90)


name = StringVar()
name2 = StringVar()
month_num = StringVar()
yr = StringVar()
#entry_box1 = Entry(root, textvariable = name, width = 25).place(x=150, y = 0)
entry_box2 = Entry(root, textvariable = name2, width = 25).place(x=350, y = 30)
entry_box3 = Entry(root, textvariable = month_num, width = 25).place(x=250, y = 60)
entry_box4 = Entry(root, textvariable = yr, width = 25).place(x=300, y = 90)


def do_it():
	#list of agencies with same one-page excel format
    cum_agencies = ["All Agencies","PACT County","Child Welfare Agencies","Sacramento County","San Francisco County","Los Angeles County"]
    #convert site name into string
    site = str(comboExample.get())
    #read table with the name given by user
    test = Table.read_table(str(name2.get()) + ".csv")
    #Popup messagebox if user mispelled agency name
    if site not in cum_agencies and site != 'State of Texas' and site not in test.group("Agency").column("Agency") and site != "Solano County" and site != "Mecklenberg County" and site != "Alameda County":
        messagebox.showinfo("Error", "The agency entered either doesn't exist or is mispelled, enter a different agency.")
    #get culmulative table for a specific site; all sites except for All Agencies, Alameda County, PACT County, and Chil Welfare Agencies can obtain a table with data from a certain month that the user provides
    if site == "State of Texas":
        crit1 = test.where("Texas", are.equal_to("Texas Sites")).where("month", are.below_or_equal_to(int(month_num.get()))).where("year", are.below_or_equal_to(int(yr.get())))
        crit2 = test.where("Texas", are.equal_to("Texas Sites")).where("month", are.above(int(month_num.get()))).where("year", are.below(int(yr.get())))
        crit = crit1.append(crit2)
    elif site == "All Agencies":
        crit = test.where("Texas", are.not_equal_to("Texas Sites"))
    elif site == "Alameda County":
        crit = test.where("county", are.equal_to("Alameda"))
    elif site == "PACT County":
        crit = test.where("pact", are.equal_to("Yes"))
    elif site == "Child Welfare Agencies":
        crit = test.where("setting", are.equal_to("child welfare"))
    elif site != "PACT County" and site != "All Agencies" and site != "Child Welfare Agencies" and site in cum_agencies or site == "Solano County" or site == "Mecklenberg County":
        word =  site[:site.index(" C")]
        crit1 = test.where("county", word).where("month", are.below_or_equal_to(int(month_num.get()))).where("year", are.below_or_equal_to(int(yr.get())))
        crit2 = test.where("county", word).where("month", are.above(int(month_num.get()))).where("year", are.below(int(yr.get())))
        crit = crit1.append(crit2)
    else:
        crit1 = test.where("Agency", site).where("month", are.below_or_equal_to(int(month_num.get()))).where("year", are.below_or_equal_to(int(yr.get())))
        crit2 = test.where("Agency", site).where("month", are.above(int(month_num.get()))).where("year", are.below(int(yr.get())))
        crit = crit1.append(crit2)
    dt = datetime.datetime.today()
    mon_lst = ['Empty', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    #Dictionary of site names and the respective date of first CSE-IT entry if provided; used to populate Date Range field
    sites = {'All Agencies': '4/10/2015', "PACT County": "12/7/2015","Child Welfare Agencies":"12/14/2015","Sacramento County":"8/7/2015","San Francisco County":"4/10/2015","Los Angeles County":"9/10/2015",
    'Solano County': '7/31/2017', 'Mecklenberg County':'12/14/2017', 'Alameda County': '3/29/2016',
    #Non-Texas
    'Bay Legal': '2/4/2016', 'Calaveras CPS': '2/21/2017', "Center for Human Services": '4/24/2018', 'CommuniCare': '8/12/2015', 
    'Community Human Service, Safe Place': '7/17/2018', 'Council for Children Rights': '12/14/2017', 'Crittenton': '10/27/2015', 'Placer County CPS': '10/19/2017', 
    'LA DMH': '3/16/2016', "Larkin": '6/11/2015', "Monarch": '5/13/2016', "Monterey FCS": '3/31/2016', "Pat’s Place Child Advocacy": '3/26/2018', "Riverside University Health System": '11/22/2016',
    "Sacramento CPS": '12/7/2015', "Sacramento Probation": '9/1/2015', "San Diego Child Welfare": '2/17/2016', "San Francisco HSA": '4/10/2015', 
    "San Mateo County CFS": '9/10/2015', "Santa Barbara Social Services": '12/30/2015', "Santa Clara DFCS": '2/8/2018', "Santa Cruz County Child Welfare": '11/17/2016', 'Santa Cruz Probation': '2/16/2016',
    "Shasta County Child Welfare": '7/6/2017', "Side by Side-North": '5/31/2018', "Side by Side-Hayward": '8/7/2015', "Solano County Behavioral Health": '7/31/2017', 
    "Street Light USA": '7/17/2018', "Turning Point Yolo": '9/25/2015', "Ventura County DCFS": '2/5/2016', "Without Permission": '2/26/2018', "Yolo Probation":  '12/17/2015', 
    "Yuba County Child Welfare": '8/28/2017', "LACY": '3/28/2016', "Present Age Ministries": '1/27/2018', "Sac City USD": '10/3/2015', "San Luis Obispo CWS": '12/14/2015', 
    "Teen Health Connection": '12/29/2017', "Youth Homes - Montana": '9/8/2017', "Children's Law Center (CLC) - Sac": '8/7/2015', 'Operation SafeHouse Riverside County': '2/10/17', 'Ventura County Probation': '5/8/2018',
    'New Jersey Court': '3/5/2018', 'Humboldt County Child Welfare': '5/17/2018', 'Plumas Child Welfare': '5/7/2018', "Contra Costa CFS":'8/4/2015', 
    #Texas
    "Atascosa County Juvenile Probation Department": "9/25/2018", "BCFS Health and Human Services": "7/13/2018", "Bexar County Probation": "12/26/2017", 
    "Calhoun County Juvenile Probation Department": "10/1/2018","Central Texas Youth Services": "9/13/2018", "Cherokee County Juvenile Services": "11/9/2018", 
    "Children's Advocacy Center for Denton County": "11/21/2018", "ChildSafe": "3/14/16", "Texas Department of Family and Protective Services": "2/20/2019", 
    "For The Silent": "11/30/2018","Garth House Mickey Mehaffy Childrens Advocacy Program": "10/31/2018", "Harris County DA": "6/27/2018",
    "Harris County Protective Services for Children and Adults": "9/4/2018", "Harvest House": "9/21/2018", "Hutchinson County Juvenile Probation": "8/14/2018", 
    "Jefferson County Juvenile Probation": "10/4/2018", "New Friends New Life": "1/2/2019", "Promise House": "6/14/2018", "Regional Victim Crisis Center": "7/5/2018", 
    "Roy Maas Youth Alternative": "12/18/2017","SAFE Alliance": "12/4/2018", "St. Jude’s Ranch for Children – Texas (SJRC)": "7/27/2018", "The Center for Success and Independence": "1/29/2019", 
    "UnBound Fort Worth": "11/13/2018", "YMCA International Services": "10/8/2018", "State of Texas": "3/14/2016", 'ARROW CHILD & FAMILY MINISTRIES':'5/7/2019',
    'Court Appointed Special Advocates of Tarrant County (CASA)': '5/7/2019'}

    #Fills in cells for the given column of the given section of the excel sheet if value is a string (used to fill out therapist name column)
    def cell(section, ind, row, col, ws, site):
        coll = col
        for a in section.column(ind):
            coll += str(row)
            ws[coll] = str(a)
            row += 1
            coll = col
        wb.save("/Volumes/Groups/WCC Research/CSEIT Site Reports/____CSE-IT Site Reports/" + str(yr.get()) + "/" + str(int(month_num.get())+1) + ". " + mon_lst[int(month_num.get())+1] + " " + str(yr.get())+ "/" + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
        #wb.save('/Users/'+getpass.getuser()+'/Desktop/sr/' + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
    #Fills in cells of the excel sheet if value is a string
    def cells(section, ind, row, col, ws, site):
        coll = col
        for a in section.column(ind):
            coll += str(row)
            ws[coll] = int(a)
            row += 1
            coll = col
        wb.save("/Volumes/Groups/WCC Research/CSEIT Site Reports/____CSE-IT Site Reports/" + str(yr.get()) + "/" + str(int(month_num.get())+1) + ". " + mon_lst[int(month_num.get())+1] + " " + str(yr.get())+ "/" + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
        #wb.save('/Users/'+getpass.getuser()+'/Desktop/sr/' + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
    #uses cells function for every column of every table    
    def cat(section, start, ws, site):
        cells(section, "count", start, 'C', ws, site)
        cells(section, "No Concern", start, 'E', ws, site)
        cells(section, "Possible Concern", start, 'G', ws, site)
        cells(section, "Clear Concern", start, 'I', ws, site)
        wb.save("/Volumes/Groups/WCC Research/CSEIT Site Reports/____CSE-IT Site Reports/" + str(yr.get()) + "/" + str(int(month_num.get())+1) + ". " + mon_lst[int(month_num.get())+1] + " " + str(yr.get())+ "/" + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
        #wb.save('/Users/'+getpass.getuser()+'/Desktop/sr/' + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
    #fills in missing categories given the full list of categories for a section
    def fill_in_row(section, full_list, col_list):
        miss_var = [i for i in full_list if i not in col_list]
        zero_lst = []
        for y in range(len(section.labels)-1):
            zero_lst += [0]
        for x in miss_var:
            section = section.with_row([x] + zero_lst)
        return section
    #returns a list of indexes to order the list of categories
    def fill_in(full_list, col_list):
        take_index = []
        for x in full_list: 
            take_index += [col_list.index(x)]
        return take_index
    #returns desired table for a specific section
    def get_table(column, table):
        no_nan = table.where(column, are.not_equal_to('nan'))
        total = no_nan.group(column)
        table = no_nan.pivot("cseit", column).join(column, total)
        return table
    #returns table with all neccessary columns
    def fill_col(section):
        columns = ['count', 'No Concern', 'Possible Concern', 'Clear Concern']
        for x in columns:
            lst = []
            if x not in list(section.labels)[1:]:
                for y in section.column(0).tolist():
                    lst += [0]
                section = section.with_columns(x, lst)
                lst = []
        return section
    #find the mean of a list of numbers
    def mean(lst): 
    	return sum(lst) / len(lst) 
    #perform package of functions in order to populate the correct cells
    def do(category, crit, lst, row, ws, site):
        table = get_table(category, crit)
        table = fill_col(table)
        table = fill_in_row(table, lst, table.column(category).tolist())
        ind = fill_in(lst, table.column(category).tolist())
        table = table.take(ind)
        cat(table, row, ws, site)
        return table
    #Orders, obtains, and populates cells relating to average age
    def age_av(crit, alp, row, ws):
        l = [x for x in crit.column("Age") if x > 0]
        m = mean(l)
        age_mean =crit.select("cseit", "Age").where("Age", are.above(0)).group("cseit", mean).with_row(["Total", m])
        lst = ["Total", "No Concern", "Possible Concern", "Clear Concern"]
        age_mean = fill_in_row(age_mean, lst, age_mean.column("cseit").tolist())
        b = fill_in(lst, age_mean.column("cseit").tolist())
        age_mean = age_mean.take(b)
        cell = ""
        index = 0
        for x in age_mean.column("Age mean"):
            cell += alp[index] + str(row)
            ws[cell] = x
            cell = ""
            index += 1
        wb.save("/Volumes/Groups/WCC Research/CSEIT Site Reports/____CSE-IT Site Reports/" + str(yr.get()) + "/" + str(int(month_num.get())+1) + ". " + mon_lst[int(month_num.get())+1] + " " + str(yr.get())+ "/" + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
        #wb.save('/Users/'+getpass.getuser()+'/Desktop/sr/' + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
    #Populates the culmulative table
    def summ(crit, ws, site):
        table = crit.group("cseit")
        lst = ['Total Number Screened', 'No Concern', 'Possible Concern', 'Clear Concern']
        cse = table.with_row(["Total Number Screened", sum(table.column("count"))])
        cse = fill_in_row(cse, lst, cse.column("cseit").tolist())
        d = fill_in(lst, cse.column("cseit").tolist())
        cse = cse.take(d)
        if site == "Solano County":
            cells(cse, "count", 21, 'C', ws, site)
        else:
            cells(cse, "count", 28, 'C', ws, site)

    days = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10:31, 11: 30, 12: 31}
    #Loads spreadsheet depending on the site
    if site == "State of Texas":
        wb = load_workbook('Blank_SOT_n.xlsx')
    elif site == "Alameda County":
        wb = load_workbook("ala_temp.xlsx")
    elif site == "Solano County":
        wb = load_workbook('solano_temp.xlsx')
    elif site == "Mecklenberg County":
        wb = load_workbook('meck_temp.xlsx')
    elif crit.column("Texas")[0] != "Texas Sites" and site not in cum_agencies:
        wb = load_workbook('n.xlsx')
    elif site in cum_agencies:
        wb = load_workbook('Blank_All.xlsx')
    else:
        wb = load_workbook('Blank_Texas_n.xlsx')
    
    #First worksheet of existing spredsheet and second worksheet if agency has a seconf worksheet
    ws = wb.worksheets[0]
    if site not in cum_agencies:
        ws1 = wb.worksheets[1]
    #Renames the worksheet
    if crit.column("Texas")[0] != "Texas Sites" or site in cum_agencies:
        ws.title = mon_lst[int(month_num.get())+1] + " " + str(yr.get())
    #fill in site name
    d = crit.sort("RecordedDate").column("RecordedDate")
    sortedArray = sorted(d, key=lambda x: datetime.datetime.strptime(x, '%m/%d/%Y %H:%M'))
    dates = [s[:s.find(" ")] for s in sortedArray]
    if site == "Solano County":
        ws['C18'] = site
        ws['C20'] = dates[len(dates)-1]
    else:
        ws['C25'] = site
        ws['C27'] = dates[len(dates)-1]
    #Finds first recorded date if not in the dictionary and poplates the date range cell
    if site not in [k for k in sites.keys()]:
        d = crit.sort("RecordedDate").column("RecordedDate")
        sortedArray = sorted(d, key=lambda x: datetime.datetime.strptime(x, '%m/%d/%Y %H:%M'))
        dates = [s[:s.find(" ")] for s in sortedArray]
        ws['C26'] = dates[0] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
    elif site != "Solano County" and site in [k for k in sites.keys()]:
        ws['C26'] = sites[site] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
    if site == "Solano County":
        ws['C19'] = sites[site] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
    #Fills out list of agency on second worksheet for the state of texas or list of therapists for every other site (beside those with only one worksheet)
    if site == "State of Texas":
        agency = get_table("Agency", crit)
        agency = fill_col(agency)
        cell(agency, "Agency", 6, 'B', ws1, site)
        cat(agency, 6, ws1, site)
    elif site not in cum_agencies and site != 'State of Texas' and site != "Mecklenberg County" and site != "Alameda County":
        therapist = crit.group("TherapistName_Email")
        cell(therapist, "TherapistName_Email", 5, 'A', ws1, site) 
        cells(therapist, "count", 5, 'B', ws1, site)
        #fills in the percentage column of the therapist table
        coll = 'C'
        row = 5
        if site == "Solano County":
            for c in crit.group("TherapistName_Email").column("TherapistName_Email"):
                coll += str(row)
                ws1[coll] = "=B" + str(row) + "/'" + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + "'!C$21:E$21"
                row += 1
                coll = 'C'
        elif crit.column("Texas")[0] != "Texas Sites" and site not in cum_agencies:
            for c in crit.group("TherapistName_Email").column("TherapistName_Email"):
                coll += str(row)
                ws1[coll] = "=B" + str(row) + "/'" + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + "'!C$28:E$28"
                row += 1
                coll = 'C'
        elif crit.column("Texas")[0] == "Texas Sites":
            for c in crit.group("TherapistName_Email").column("TherapistName_Email"):
                coll += str(row)
                ws1[coll] = "=B" + str(row) + "/'Cumulative Report'!C$28:E$28"
                row += 1
                coll = 'C'
    #Populates the gender id, race, average age, and age category tables
    gend_id = ['male', 'female', 'other', 'Unknown']
    race = ['African American or Black', 'Asian', 'Native American', 'Hispanic or Latino', 'White or Caucasian', 'Unknown', 'Multiracial', 'Other']
    alph = ['C', 'D', 'E', 'F']
    age_cat = ['<10 yrs','10-11 years', '12-13 years', '14-15 years','16-17 years','18-19 years', '20-21 years','22-23 years','24+ years', 'Unknown']
    summ(crit, ws, site)
    do("SexMaleFemale", crit, gend_id, 39, ws, site)
    do("Race", crit, race, 50, ws, site)
    age_av(crit, alph, 65, ws)
    do("agecat", crit, age_cat, 73, ws, site)
    #Fills in sexual orientation, gender expression, disability, and forced labor table for non-texas sites
    if crit.column("Texas")[0] != "Texas Sites":
        sex_or = ['heterosexual','gay', 'bisexual', 'lesbian','unknown', 'client unsure', 'decline to answer']
        gend_exp = ['yes','no', 'unknown']
        disa = ['None','Physical disability', 'Intellectual disability', 'Both','Unknown']
        labor = ['Yes', 'No', 'Unknown']
        do("SexualOrientation", crit, sex_or, 90, ws, site)
        do("GenderExpression", crit, gend_exp, 104, ws, site)
        do("disability", crit, disa, 115, ws, site)
        do("forcedlabor", crit, labor, 128, ws, site)
        #if site has data for "Yes" on forced labor, fill out table for sex, race, and age category for forced labor
        if site in cum_agencies or int(do("forcedlabor", crit, labor, 128, ws, site).where("forcedlabor", are.equal_to("Yes")).column("count").item(0)):
            total = crit.where("forcedlabor", are.equal_to("Yes"))
            do("SexMaleFemale", total, gend_id, 139, ws, site)
            do("Race", total, race, 150, ws, site)
            do("agecat", total, age_cat, 166, ws, site)
    #Solano County Agency table
    if site == "Solano County":
        ag = crit.group("Agency")
        cell(ag, "Agency", 28, 'B', ws, site) 
        cell(ag, "count", 28, 'C', ws, site)
    #Mecklenberg County Client ID table on second worksheet
    if site == "Mecklenberg County":
        coll = 'C'
        row = 5
        clients = crit.where("cseit", are.equal_to("Clear Concern")).group("ClientID")
        cell(clients, "ClientID", 5, 'A', ws1, site) 
        cells(clients, "count", 5, 'B', ws1, site)
        for c in crit.where("cseit", are.equal_to("Clear Concern")).group("ClientID").column("ClientID"):
            coll += str(row)
            ws1[coll] = "=B" + str(row) + "/'" + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + "'!C$30"
            row += 1
            coll = 'C'
    #Alameda special tables
    if site == "Alameda County":
        ag_con = get_table("Agency", crit)
        ag_con = fill_col(ag_con)
        cat(ag_con, 33, ws1, site)
        cell(ag_con, "Agency", 33, 'B', ws1, site)
    
        def alameda(site, row, ws1, ala, row1):
            where = crit.where("Agency", site)
            alp = ['C', 'E', 'G', 'I']
            age_cat = ['<10 yrs','10-11 years', '12-13 years', '14-15 years','16-17 years','18-19 years', '20-21 years','22-23 years','24+ years', 'Unknown']
            do("agecat", where, age_cat, row, ws1, ala)
            age_av(where, alp, row1, ws1)
        #Age category tables for Alameda County agencies
        alameda("Bay Legal", 61, ws1, site,60)
        alameda("Side by Side-Hayward", 78, ws1, site,77)
        alameda("DreamCatcher Alameda", 95, ws1, site,94)
        alameda("WestCoast C-Change", 112, ws1, site,111)
        alameda("WestCoast Pre- and PostDocs", 129, ws1, site,128)
        alameda("WestCoast OPT", 146, ws1, site,145)
        alameda("WestCoast STAT", 163, ws1, site, 162)
        alameda("Catholic Charities", 180, ws1, site,179)
        alameda("WestCoast Catch-21", 197, ws1, site,196)
        alameda("A Better Way Inc", 214, ws1, site,213)
        alameda("WestCoast Children's Clinic Unknown", 231, ws1, site,230)

#Texas
    if crit.column("Texas")[0] == "Texas Sites":
    	#different Texas wokrsheets
        cum_graph = wb.worksheets[7]
        mon_graph = wb.worksheets[9]
        blank = wb.worksheets[3]
        monthly_rep = wb.worksheets[2]
        m_therapist = wb.worksheets[4]
        cum_data = wb.worksheets[5]
        data_cum = wb.worksheets[6]
        graph_cum = wb.worksheets[8]
        #fills in data range for worksheets with graphs
        if site not in [k for k in sites.keys()]:
            d = crit.sort("RecordedDate").column("RecordedDate")
            sortedArray = sorted(d, key=lambda x: datetime.datetime.strptime(x, '%m/%d/%Y %H:%M'))
            dates = [s[:s.find(" ")] for s in sortedArray]
            cum_graph['A28'] = "Data Range: " + dates[0] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
            graph_cum['A28'] = "Data Range: " + dates[0] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
            mon_graph['A28'] = "Data Range: " + str(month_num.get()) + "/1/" + str(yr.get()) + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
        else:
            graph_cum['A28'] = "Data Range: " + sites[site] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
            cum_graph['A28'] = "Data Range: " + sites[site] + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
            mon_graph['A28'] = "Data Range: " + str(month_num.get()) + "/1/" + str(yr.get()) + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
        if site != "State of Texas":
            cum_graph['A17'] = site
            graph_cum['A17'] = site
            mon_graph['A17'] = site
        #Gets monthly data
        if site == "State of Texas":
            crit2 = test.where("Texas", are.equal_to("Texas Sites")).where("month", are.equal_to(int(month_num.get()))).where('year', are.equal_to(int(yr.get())))
        else:
            crit2 = test.where("Agency", site).where("month", are.equal_to(int(month_num.get()))).where('year', are.equal_to(int(yr.get())))
        #remove certain worksheets if there is not new data for the month     
        if len(crit2.column("Agency")) == 0:
            blank['B25'] = "No new CSE-IT data reported for " + mon_lst[int(month_num.get())] + " 2019."
            wb.remove(mon_graph)
            wb.remove(cum_graph)
            wb.remove(cum_data)
            wb.remove(m_therapist)
            wb.remove(monthly_rep)
        else:
        #Fills monthly Agencies table for Texas and monthly therapists table for other sites
            monthly_rep['C25'] = site
            monthly_rep['C26'] = str(month_num.get()) + "/1/" + str(yr.get()) + " - " + str(month_num.get()) + "/" + str(days[int(month_num.get())]) + "/" + str(yr.get())
            monthly_rep['C27'] = dates[len(dates)-1]
            if site == "State of Texas":
                agency = get_table("Agency", crit2)
                agency = fill_col(agency)
                cell(agency, "Agency", 6, 'B', m_therapist, site)
                cat(agency, 6, m_therapist, site)
            else:
                therapist = crit2.group("TherapistName_Email")
                cell(therapist, "TherapistName_Email", 5, 'A', m_therapist, site) 
                cells(therapist, "count", 5, 'B', m_therapist, site)
                coll = 'C'
                row = 5
                for c in crit2.group("TherapistName_Email").column("TherapistName_Email"):
                    coll += str(row)
                    m_therapist[coll] = "=B" + str(row) + "/'Monthly Report'!C$28:E$28"
                    row += 1
                    coll = 'C'
            #Fills culmulative, gender, race, average age, and age category tables for the month
            summ(crit2, monthly_rep, site)
            do("SexMaleFemale", crit2, gend_id, 39, monthly_rep, site)
            do("Race", crit2, race, 50, monthly_rep, site)
            age_av(crit2, alph, 65, monthly_rep)
            do("agecat", crit2, age_cat, 73, monthly_rep, site)

            wb.remove(graph_cum)
            wb.remove(data_cum)
            wb.remove(blank)
    wb.save("/Volumes/Groups/WCC Research/CSEIT Site Reports/____CSE-IT Site Reports/" + str(yr.get()) + "/" + str(int(month_num.get())+1) + ". " + mon_lst[int(month_num.get())+1] + " " + str(yr.get())+ "/" + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')
    #wb.save('/Users/'+getpass.getuser()+'/Desktop/sr/' + site + " CSE-IT Site Report " + mon_lst[int(month_num.get())+1] + " " + str(yr.get()) + '.xlsx')

work = Button(root, text = "Submit", width = 30, command = do_it).place(x = 150, y = 120)    
root.mainloop()
